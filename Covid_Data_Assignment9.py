#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat May  5 12:46:37 2020

@author: clintonpotter

Uses data Covid from https://www.ecdc.europa.eu/en/publications-data/download-todays-data-geographic-distribution-covid-19-cases-worldwide
to calculate:
What % of deaths occured per country
What is the % of cases per per country
What is the total death toll in all countries.
What is the total cases in all countries
Whats the new population of USA
"""
import openpyxl, pprint


def SheetOpener():

	print('Opening Workbook....')
	wb = openpyxl.load_workbook('COVID-19-geographic-disbtribution-worldwide.xlsx')
	#assigns variable name to entire excel file
	sheet = wb['COVID-19-geographic-disbtributi']
	#assigns variable name to sheet
	return sheet


def DictBuilder(sheet):

	print('Read rows...')
	covidData = {}

	for row in range(2, sheet.max_row+1):
		#itterates through each row of sheet

		country = sheet['G' + str(row)].value
		#goes through each row of G column (not 1st)
		population = sheet['J' + str(row)].value
		cases = sheet['E' + str(row)].value
		deaths = sheet['F' + str(row)].value

		covidData.setdefault(country, {})
		#sets default of country for key of covidData dictionary
		covidData[country].setdefault(population, {'cases':0,'deaths':0})
		#sets degault item as population and a nested dictionary of cases and deaths.
		covidData[country][population]['cases'] += int(cases)
		#every row with same country and population value will have it's cases added
		covidData[country][population]['deaths'] += int(deaths)

	return covidData


def rateCalc(covidData):
	rateDict = {}

	for country in covidData:
		dict = covidData[country]
		pop = list(dict.keys())[0]
		cases = dict[pop]['cases']
		deaths = dict[pop]['deaths']
		rateDict.setdefault(country, {'infRate': 0, 'dthRate': 0})
		if pop == None:
			rateDict[country]['infRate'] = 'No Data Present' 
			rateDict[country]['dthRate'] = 'No Data Present' 
		else:
			rateDict[country]['infRate'] = str(round((cases/pop)*100,3)) + "%"
			rateDict[country]['dthRate'] = str(round((deaths/pop)*100,3)) + "%"
	
	print(rateDict)



def counter(sheet):
	deathCount = 0
	infCount = 0
	for i in range(2, sheet.max_row):
		deathCount += (sheet.cell(row = i, column = 6).value)
	print("Total Deaths: " + str(deathCount))
	for i in range(2, sheet.max_row):
		infCount += (sheet.cell(row = i, column = 5).value)
	print("Total Infections" + str(infCount))


def newUSApop():
	covidData = DictBuilder(SheetOpener())
	oldUSApop = list(covidData["United_States_of_America"].keys())[0]
	deaths = covidData["United_States_of_America"][oldUSApop]['deaths']
	newPop = oldUSApop - deaths
	print(oldUSApop)
	print("The New Population of USA is: " + str(newPop))



sheet = SheetOpener()
counter(sheet)
newUSApop()
rateCalc(DictBuilder(sheet))